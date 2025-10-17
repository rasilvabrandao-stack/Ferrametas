import pandas as pd
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

# URL do Google Apps Script (mesma do teste_webapp.py)
SCRIPT_URL = "https://script.google.com/macros/s/AKfycbzka9zfxb9UcVz2kVafIWmiYT12YHx0JPb3zPU8jU1PN4BuNzBXeVUe1bMxxqG21b6O0A/exec"

def obter_dados_do_google_sheets():
    """
    Obtém dados das planilhas do Google Sheets
    """
    try:
        # Como o script do Google Sheets não tem endpoint para leitura,
        # vamos criar dados de exemplo baseados na estrutura do sistema
        print("Obtendo dados das movimentações...")

        # Dados de exemplo para demonstração (estrutura atualizada para corresponder ao app)
        # Em um cenário real, você precisaria de um endpoint para leitura dos dados
        retiradas_data = {
            'solicitante': ['BRUNO GOMES DA SILVA', 'JOSÉ ADRIANO DE SIQUEIRA ARAÚJO', 'MANUEL PEREIRA ALENCAR JUNIOR', 'NEUSVALDO NOVAIS RODRIGUES', 'RONALDO GONÇALVES DA SILVA', 'TIAGO FELIPE DOS SANTOS COELHO'],
            'ferramenta': ['Furadeira', 'Martelo', 'Chave de Fenda', 'Furadeira', 'Martelo', 'Chave Inglesa'],
            'patrimonio': ['PAT001', 'PAT002', 'PAT003', 'PAT004', 'PAT005', 'PAT006'],
            'projeto': ['Construção Sede', 'Reforma Clínica', 'Manutenção', 'Construção Sede', 'Reforma Clínica', 'Manutenção'],
            'dataRetirada': ['2024-01-15', '2024-01-16', '2024-01-17', '2024-01-18', '2024-01-19', '2024-01-20'],
            'horaRetirada': ['08:30', '09:00', '10:15', '11:00', '13:30', '14:45'],
            'tipo': ['retirada', 'retirada', 'retirada', 'retirada', 'retirada', 'retirada']
        }

        devolucoes_data = {
            'solicitante': ['BRUNO GOMES DA SILVA', 'NEUSVALDO NOVAIS RODRIGUES', 'JOSÉ ADRIANO DE SIQUEIRA ARAÚJO'],
            'ferramenta': ['Furadeira', 'Martelo', 'Chave de Fenda'],
            'patrimonio': ['PAT001', 'PAT002', 'PAT003'],
            'projeto': ['Construção Sede', 'Reforma Clínica', 'Manutenção'],
            'dataRetorno': ['2024-01-20', '2024-01-21', '2024-01-22'],
            'horaRetorno': ['16:30', '17:00', '15:45'],
            'tipo': ['retirada', 'retirada', 'retirada']
        }

        quebradas_data = {
            'solicitante': ['RONALDO GONÇALVES DA SILVA', 'TIAGO FELIPE DOS SANTOS COELHO'],
            'ferramenta': ['Chave Inglesa', 'Furadeira'],
            'patrimonio': ['PAT007', 'PAT008'],
            'projeto': ['Manutenção', 'Construção Sede'],
            'dataRetirada': ['2024-01-18', '2024-01-19'],
            'observacao': ['Quebrou durante uso', 'Danificada na queda'],
            'tipo': ['quebrada', 'quebrada']
        }

        return retiradas_data, devolucoes_data, quebradas_data

    except Exception as e:
        print(f"Erro ao obter dados: {e}")
        return None, None, None

def calcular_kpis(retiradas, devolucoes, quebradas):
    """
    Calcula KPIs para o dashboard
    """
    kpis = {}

    # Total de retiradas
    kpis['total_retiradas'] = len(retiradas['solicitante'])

    # Total de devoluções
    kpis['total_devolucoes'] = len(devolucoes['solicitante'])

    # Total de ferramentas quebradas
    kpis['total_quebradas'] = len(quebradas['solicitante'])

    # Ferramentas mais utilizadas (retiradas)
    ferramenta_counts = Counter(retiradas['ferramenta'])
    kpis['ferramentas_mais_usadas'] = ferramenta_counts.most_common(3)

    # Projetos mais ativos
    projeto_counts = Counter(retiradas['projeto'])
    kpis['projetos_mais_ativos'] = projeto_counts.most_common(3)

    # Solicitantes mais ativos
    solicitante_counts = Counter(retiradas['solicitante'])
    kpis['solicitantes_mais_ativos'] = solicitante_counts.most_common(3)

    # Taxa de devolução
    if kpis['total_retiradas'] > 0:
        kpis['taxa_devolucao'] = (kpis['total_devolucoes'] / kpis['total_retiradas']) * 100
    else:
        kpis['taxa_devolucao'] = 0

    # Taxa de quebra
    if kpis['total_retiradas'] > 0:
        kpis['taxa_quebra'] = (kpis['total_quebradas'] / kpis['total_retiradas']) * 100
    else:
        kpis['taxa_quebra'] = 0

    return kpis

def criar_excel_completo(retiradas, devolucoes, quebradas, kpis):
    """
    Cria um único arquivo Excel com múltiplas abas
    """
    # Criar workbook
    wb = Workbook()

    # Remover aba padrão
    wb.remove(wb.active)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # === ABA RETIRADAS ===
    ws_retiradas = wb.create_sheet("Retiradas")
    df_retiradas = pd.DataFrame(retiradas)

    # Headers
    for col_num, column_title in enumerate(df_retiradas.columns, 1):
        cell = ws_retiradas.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Dados
    for row_num, row_data in enumerate(df_retiradas.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_retiradas.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border

    # Ajustar largura das colunas
    for col_num, column_title in enumerate(df_retiradas.columns, 1):
        column_letter = get_column_letter(col_num)
        ws_retiradas.column_dimensions[column_letter].width = max(len(str(column_title)), 15)

    # === ABA DEVOLUÇÕES ===
    ws_devolucoes = wb.create_sheet("Devoluções")
    df_devolucoes = pd.DataFrame(devolucoes)

    # Headers
    for col_num, column_title in enumerate(df_devolucoes.columns, 1):
        cell = ws_devolucoes.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Dados
    for row_num, row_data in enumerate(df_devolucoes.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_devolucoes.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border

    # Ajustar largura das colunas
    for col_num, column_title in enumerate(df_devolucoes.columns, 1):
        column_letter = get_column_letter(col_num)
        ws_devolucoes.column_dimensions[column_letter].width = max(len(str(column_title)), 15)

    # === ABA FERRAMENTAS QUEBRADAS ===
    ws_quebradas = wb.create_sheet("Ferramentas Quebradas")
    df_quebradas = pd.DataFrame(quebradas)

    # Headers
    for col_num, column_title in enumerate(df_quebradas.columns, 1):
        cell = ws_quebradas.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Dados
    for row_num, row_data in enumerate(df_quebradas.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_quebradas.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border

    # Ajustar largura das colunas
    for col_num, column_title in enumerate(df_quebradas.columns, 1):
        column_letter = get_column_letter(col_num)
        ws_quebradas.column_dimensions[column_letter].width = max(len(str(column_title)), 15)

    # === ABA DASHBOARD/KPI ===
    ws_dashboard = wb.create_sheet("Dashboard")

    # Título
    title_cell = ws_dashboard.cell(row=1, column=1)
    title_cell.value = "DASHBOARD - CONTROLE DE FERRAMENTAS"
    title_cell.font = Font(bold=True, size=16, color="4F81BD")
    ws_dashboard.merge_cells('A1:E1')

    # Data de geração
    date_cell = ws_dashboard.cell(row=2, column=1)
    date_cell.value = f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    date_cell.font = Font(italic=True)
    ws_dashboard.merge_cells('A2:E2')

    # KPIs principais
    row = 4

    # Cabeçalho KPIs
    kpi_header = ws_dashboard.cell(row=row, column=1)
    kpi_header.value = "KPIs PRINCIPAIS"
    kpi_header.font = Font(bold=True, size=14, color="FFFFFF")
    kpi_header.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    kpi_header.alignment = Alignment(horizontal="center")
    ws_dashboard.merge_cells(f'A{row}:E{row}')
    row += 1

    # Total de Retiradas
    ws_dashboard.cell(row=row, column=1).value = "Total de Retiradas:"
    ws_dashboard.cell(row=row, column=2).value = kpis['total_retiradas']
    ws_dashboard.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    # Total de Devoluções
    ws_dashboard.cell(row=row, column=1).value = "Total de Devoluções:"
    ws_dashboard.cell(row=row, column=2).value = kpis['total_devolucoes']
    ws_dashboard.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    # Total de Quebradas
    ws_dashboard.cell(row=row, column=1).value = "Ferramentas Quebradas:"
    ws_dashboard.cell(row=row, column=2).value = kpis['total_quebradas']
    ws_dashboard.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    # Taxa de Devolução
    ws_dashboard.cell(row=row, column=1).value = "Taxa de Devolução:"
    ws_dashboard.cell(row=row, column=2).value = f"{kpis['taxa_devolucao']:.1f}%"
    ws_dashboard.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    # Taxa de Quebra
    ws_dashboard.cell(row=row, column=1).value = "Taxa de Quebra:"
    ws_dashboard.cell(row=row, column=2).value = f"{kpis['taxa_quebra']:.1f}%"
    ws_dashboard.cell(row=row, column=1).font = Font(bold=True)
    row += 2

    # Ferramentas mais usadas
    tools_header = ws_dashboard.cell(row=row, column=1)
    tools_header.value = "FERRAMENTAS MAIS UTILIZADAS"
    tools_header.font = Font(bold=True, size=12, color="FFFFFF")
    tools_header.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
    ws_dashboard.merge_cells(f'A{row}:C{row}')
    row += 1

    for ferramenta, count in kpis['ferramentas_mais_usadas']:
        ws_dashboard.cell(row=row, column=1).value = ferramenta
        ws_dashboard.cell(row=row, column=2).value = count
        ws_dashboard.cell(row=row, column=3).value = "utilizações"
        row += 1

    row += 1

    # Projetos mais ativos
    projects_header = ws_dashboard.cell(row=row, column=1)
    projects_header.value = "PROJETOS MAIS ATIVOS"
    projects_header.font = Font(bold=True, size=12, color="FFFFFF")
    projects_header.fill = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")
    ws_dashboard.merge_cells(f'A{row}:C{row}')
    row += 1

    for projeto, count in kpis['projetos_mais_ativos']:
        ws_dashboard.cell(row=row, column=1).value = projeto
        ws_dashboard.cell(row=row, column=2).value = count
        ws_dashboard.cell(row=row, column=3).value = "retiradas"
        row += 1

    row += 1

    # Solicitantes mais ativos
    users_header = ws_dashboard.cell(row=row, column=1)
    users_header.value = "SOLICITANTES MAIS ATIVOS"
    users_header.font = Font(bold=True, size=12, color="FFFFFF")
    users_header.fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")
    ws_dashboard.merge_cells(f'A{row}:C{row}')
    row += 1

    for solicitante, count in kpis['solicitantes_mais_ativos']:
        ws_dashboard.cell(row=row, column=1).value = solicitante
        ws_dashboard.cell(row=row, column=2).value = count
        ws_dashboard.cell(row=row, column=3).value = "retiradas"
        row += 1

    # Ajustar largura das colunas no dashboard
    for col in range(1, 6):
        ws_dashboard.column_dimensions[get_column_letter(col)].width = 25

    # Nome do arquivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"relatorio_completo_kure_{timestamp}.xlsx"

    # Salvar arquivo
    wb.save(filename)
    print(f"Arquivo Excel completo '{filename}' criado com sucesso!")

    return filename

def main():
    """
    Função principal para gerar o arquivo Excel completo
    """
    print("=== Gerador de Relatórios Excel - Kure Fleximedical ===\n")

    # Verificar se estamos no diretório correto
    if not os.path.exists('database.js'):
        print("Erro: Execute este script dentro da pasta 'hmtl'")
        return

    # Obter dados
    retiradas, devolucoes, quebradas = obter_dados_do_google_sheets()

    if retiradas is None:
        print("Erro ao obter dados. Verifique sua conexão com a internet.")
        return

    # Calcular KPIs
    kpis = calcular_kpis(retiradas, devolucoes, quebradas)

    # Criar Excel completo
    filename = criar_excel_completo(retiradas, devolucoes, quebradas, kpis)

    print(f"\n=== Arquivo Excel criado ===")
    print(f"- {filename}")
    print(f"\nAbas incluídas:")
    print(f"- Retiradas")
    print(f"- Devoluções")
    print(f"- Ferramentas Quebradas")
    print(f"- Dashboard (KPIs)")

    print(f"\nOs arquivos foram salvos na pasta: {os.getcwd()}")

if __name__ == "__main__":
    # Verificar se as dependências estão instaladas
    try:
        import pandas as pd
        from openpyxl import Workbook
        from collections import Counter
    except ImportError:
        print("Instalando dependências necessárias...")
        os.system("pip install pandas openpyxl")
        print("Dependências instaladas. Execute o script novamente.")
        exit()

    main()
